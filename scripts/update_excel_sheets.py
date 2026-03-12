#!/usr/bin/env python3
"""
Update both Excel tracking sheets with current test counts and implementation status.

Sources of truth:
  - AC tag counts: grep across backend/tests/*.py for '# AC: UIB-XXX'
  - Implementation status: derived from code evidence (JIRA_TRACKER.md)
  - Gap notes: from VALIDATION_GAP_ANALYSIS.md

Usage:
    python3 scripts/update_excel_sheets.py
"""

import re
import glob
from collections import Counter
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# ── 1. Count AC tags from all test files ─────────────────────────────────────

def count_test_tags() -> dict[str, int]:
    counts: Counter = Counter()
    for f in glob.glob(str(PROJECT_ROOT / "backend" / "tests" / "*.py")):
        with open(f) as fh:
            for line in fh:
                for m in re.findall(r"# AC: (UIB-\d+)", line):
                    counts[m] += 1
    return dict(counts)


# ── 2. Story-level data (from JIRA_TRACKER + code evidence) ─────────────────

# Implementation status + gap notes per story
STORY_STATUS: dict[str, tuple[str, str]] = {
    # (Implementation Status, Gap/Notes)
    "UIB-1":   ("Partial",  "Frontend exists as standalone React app, not embeddable widget. CORS tests tagged."),
    "UIB-3":   ("Partial",  "JWT auth exists with test tags. No SSO/SAML/LDAP integration."),
    "UIB-10":  ("Done",     "JWT carries role field; persona derived in chat_router. 3 new tests (audit #3)."),
    "UIB-14":  ("Done",     "/health + /auth/token + TokenResponse verified. 3 new tests (audit #3)."),
    "UIB-18":  ("Done",     "Agent chat() accepts free-text input. Retry, error handling tested."),
    "UIB-23":  ("Done",     "semantic_search tool queries FAISS/Azure from ingested docs. Doc store + RAG config tests."),
    "UIB-27":  ("Done",     "_filter_by_role() in tools.py + _filter_sources_by_role() in agent.py."),
    "UIB-31":  ("Done",     "RBAC filter + safe message + source format. 4 tests (audit #3)."),
    "UIB-35":  ("Done",     "Citation pipeline: extract → filter → enrich. display_name + uploaded_at."),
    "UIB-40":  ("Done",     "Azure OData filter + FAISS client-side filter. 5 tests (audit #3)."),
    "UIB-44":  ("Done",     "Neutral fallback clears sources array."),
    "UIB-48":  ("Pending",  "Not implemented. No configurable next-step suggestions (Gap M7)."),
    "UIB-52":  ("Done",     "RBAC filter prevents restricted doc names in citations."),
    "UIB-57":  ("Pending",  "No version governance code."),
    "UIB-61":  ("Pending",  "No version/draft/approved status fields."),
    "UIB-65":  ("Pending",  "No draft/archive/superseded filtering."),
    "UIB-69":  ("Pending",  "Citations include filename+page but no version/effective-date."),
    "UIB-75":  ("Partial",  "RAG retriever returns multi-doc. 1 new test (audit #3)."),
    "UIB-79":  ("Partial",  "LLM consolidates from multiple observations. 2 new tests (audit #3)."),
    "UIB-83":  ("Partial",  "_extract_sources() supports multiple sources. 3 tests."),
    "UIB-88":  ("Pending",  "No disambiguation for large result sets."),
    "UIB-93":  ("Done",     "ConversationBufferWindowMemory + session TTL + isolation. 7 tags incl. TTL tests."),
    "UIB-98":  ("Done",     "clear_session() + /chat/clear endpoint. 3 new tests (audit #3)."),
    "UIB-103": ("Done",     "_is_session_expired() + _cleanup_expired_sessions(). 2 new tests (audit #3)."),
    "UIB-109": ("Done",     "get_or_create_session() creates fresh memory. 2 new tests (audit #3)."),
    "UIB-113": ("Partial",  "Prompt has CLARIFICATION section. 2 new tests (audit #3)."),
    "UIB-117": ("Partial",  "Relies on LLM prompt instruction; no deterministic detection. 1 test."),
    "UIB-123": ("Done",     "_is_fallback_response() detects 5 fallback phrases."),
    "UIB-126": ("Done",     "Prompt enforces safe fallback; _direct_faiss_search() last resort."),
    "UIB-131": ("Done",     "save_escalation() writes to DynamoDB hm-escalations."),
    "UIB-135": ("Done",     "_notify_slack() webhook + /admin/escalations endpoint."),
    "UIB-140": ("Done",     "FORM GUIDANCE prompt section. 6/6 AC done."),
    "UIB-143": ("Partial",  "FORM LOCATION GUIDANCE prompt. 5/6 AC done (AC4 blocked on UC-16)."),
    "UIB-148": ("Done",     "workflow_guard.py with regex patterns. detect_workflow_attempt() wired in endpoint."),
    "UIB-152": ("Done",     "workflow_refusal_message returned in chat_router."),
    "UIB-157": ("Done",     "IRRELEVANT QUERIES prompt section with off-topic examples. 2 tests."),
    "UIB-161": ("Done",     "Polite decline template in prompt. 1 test."),
    "UIB-166": ("Done",     "guardrails.py Layer1 regex + Layer2 LLM classification."),
    "UIB-170": ("Done",     "GuardrailViolation returns 400 with safe message."),
    "UIB-174": ("Done",     "Guardrail violations logged + escalated."),
    "UIB-179": ("Done",     "feedback_store.py + POST /feedback endpoint."),
    "UIB-183": ("Done",     "Optional comment field in feedback model."),
    "UIB-188": ("Pending",  "No admin UI for role/persona management."),
    "UIB-192": ("Pending",  "No connector configuration UI."),
    "UIB-196": ("Pending",  "No indexing schedule or sync controls."),
    "UIB-200": ("Pending",  "No admin UI for session/experience settings."),
    "UIB-204": ("Pending",  "No admin UI for fallback message customization."),
    "UIB-208": ("Partial",  "Backend API /admin/escalations exists + Slack webhook. No admin UI."),
    "UIB-212": ("Pending",  "Theme hardcoded in frontend; no admin branding UI."),
    "UIB-216": ("Pending",  "No analytics dashboard."),
    "UIB-482": ("Pending",  "No file format configuration UI."),
    "UIB-487": ("Pending",  "No connector governance UI."),
}

# ── 3. UC-level summary for Validation sheet ─────────────────────────────────

# Row 2..19 in Validation sheet, keyed by row number
UC_VALIDATION: dict[int, tuple[str, str, str]] = {
    # row: (Implementation Status, Validation Gaps, Observations)
    2: (
        "Partially implemented. SSO not integrated; standalone React frontend with JWT auth. CORS configured. 7+4 tests.",
        "SSO/SAML integration pending. Widget embedding into university portal not done.",
        "UC-01 stories UIB-1/UIB-3 partial (no SSO). UIB-10/UIB-14 fully done with 3+3 tests."
    ),
    3: (
        "Fully implemented and tested. RBAC citation filter, citation enrichment, fallback handling all verified. 76 tests across 7 stories.",
        "M2: Response formatting (prompt tuning). M3: Document audit log not implemented. M6: Outage fallback message ambiguous.",
        "Critical gaps C1-C4 all fixed (RBAC citation filter). M1 citation enrichment done. 19 doc-store tests, 16 citation tests, 10 RBAC tests."
    ),
    # Row 4 is empty/None in original — skip
    5: (
        "Fully implemented. RBAC search filter + citation filter + neutral fallback. 10 tests across 3 stories.",
        "M7 (UIB-48): Configurable next-step suggestions not implemented. M8 fixed.",
        "OData filter for Azure, client-side for FAISS. Neutral response clears sources. UIB-48 pending."
    ),
    6: (
        "Not implemented. Version governance not in scope.",
        "All 4 stories (UIB-57/61/65/69) pending. No version/draft/approved fields.",
        "UC-04 deferred — no observations recorded."
    ),
    7: (
        "Partially implemented. RAG returns multi-doc results; no explicit aggregation UX. 6 tests.",
        "UIB-88 (disambiguation for large result sets) not implemented.",
        "retriever_top_k=3 returns multiple docs. LLM consolidates. 1+2+3 tests for UIB-75/79/83."
    ),
    8: (
        "Fully implemented and tested. Session memory with TTL, cleanup, isolation. 12 tests.",
        "Point 5 (encryption at rest) is infrastructure concern.",
        "ConversationBufferWindowMemory(k=10) + TTL aligned with JWT. 7+3+2 tests for UIB-93/98. Session clear verified."
    ),
    9: (
        "Fully implemented and tested. Session expiry eviction + fresh context on new session. 4 tests.",
        "None — all ACs covered.",
        "UIB-103: _cleanup_expired_sessions(). UIB-109: _create_executor for fresh sessions. 2+2 tests."
    ),
    10: (
        "Partially implemented. Prompt-based clarification; no deterministic detection. 3 tests.",
        "Relies on LLM following prompt instructions. No programmatic ambiguity detection.",
        "UIB-113: clarification section in prompt (2 tests). UIB-117: prompt instructs single question (1 test)."
    ),
    11: (
        "Fully implemented and tested. Fallback detection + safe response + no sources. 11 tests.",
        "None — all ACs covered.",
        "UIB-123: _is_fallback_response() with 5 phrases (8 tests). UIB-126: safe fallback enforced (3 tests)."
    ),
    12: (
        "Fully implemented and tested. DynamoDB escalation store + Slack webhook. 13 tests.",
        "None — all ACs covered.",
        "UIB-131: save_escalation() (5 tests). UIB-135: Slack notify + admin endpoint (6 tests). Config tests (2)."
    ),
    13: (
        "Fully implemented. Form guidance + location guidance in prompt. 11 tests.",
        "UIB-143 AC4 blocked on UC-16 (admin-configurable form URLs).",
        "UIB-140: 6/6 AC done (6 tests). UIB-143: 5/6 AC done (5 tests)."
    ),
    14: (
        "Fully implemented and tested. Workflow guard with regex patterns + endpoint interception. 21 tests.",
        "None — all ACs covered.",
        "UIB-148: detect_workflow_attempt() wired before agent (16 tests). UIB-152: refusal message (3 tests). Config (2 tests)."
    ),
    15: (
        "Fully implemented and tested. Prompt-based irrelevant query handling. 3 tests.",
        "None — relies on LLM prompt following.",
        "UIB-157: off-topic examples in prompt (2 tests). UIB-161: polite decline (1 test)."
    ),
    16: (
        "Fully implemented and tested. Layer1 regex + Layer2 LLM guardrails. 27 tests.",
        "None — all ACs covered.",
        "UIB-166: 14 tests. UIB-170: 12 tests. UIB-174: 1 test. Two-layer defense-in-depth."
    ),
    17: (
        "Fully implemented and tested. Feedback store + rating + comments. 12 tests.",
        "None — all ACs covered.",
        "UIB-179: 10 tests. UIB-183: 2 tests. DynamoDB hm-feedback table."
    ),
    18: (
        "Mostly not implemented. Only backend escalation API + Slack webhook exist. No admin UI.",
        "9 of 10 features pending. Role management, connectors, indexing, session controls, branding, analytics — all need admin UI.",
        "UIB-208 partial (backend API exists). All other UC-16 stories pending. Blocked on admin frontend."
    ),
    19: (
        "Out of scope. UC-18 added post-original scope. No spec or implementation.",
        "Not in original UC-01 to UC-16 scope.",
        "Jira epic UIB-513. RAG quality issues apply generally, not Academic Search-specific."
    ),
}


def update_requirements_xlsx(tag_counts: dict[str, int]):
    """Update INTERNALCHATBOT_REQUIREMENTS_v5.xlsx — all 3 sheets."""
    path = PROJECT_ROOT / "INTERNALCHATBOT_REQUIREMENTS_v5.xlsx"
    wb = openpyxl.load_workbook(str(path))

    # ── Sheet: Story Tracker ──────────────────────────────────────────────
    ws = wb["Story Tracker"]
    for row in range(2, ws.max_row + 1):
        story_id = ws.cell(row, 3).value  # Column C = Story ID
        if not story_id:
            continue
        # Update Tests Tagged (column H = 8)
        count = tag_counts.get(story_id, 0)
        ws.cell(row, 8).value = count if count > 0 else ws.cell(row, 8).value

        # Update Implementation Status (column G = 7) and Gap/Notes (column I = 9)
        if story_id in STORY_STATUS:
            status, notes = STORY_STATUS[story_id]
            ws.cell(row, 7).value = status
            ws.cell(row, 9).value = notes

    # ── Sheet: Internal Chatbot Requirements (UC-level) ───────────────────
    ws_uc = wb["Internal Chatbot Requirements"]
    # Map UC IDs to aggregated status
    UC_STATUS: dict[str, tuple[str, str]] = {
        "UC-01": ("Partial", "SSO not integrated. UIB-10/14 done. UIB-1/3 partial (no widget embed, no SAML)."),
        "UC-02": ("Done", "All critical gaps fixed. RBAC citation filter, enrichment, fallback. 76 tests. M2/M3/M6 minor gaps remain."),
        "UC-03": ("Done", "RBAC search + citation filter + neutral fallback. UIB-48 (next steps) pending."),
        "UC-04": ("Pending", "Version governance not implemented."),
        "UC-05": ("Partial", "RAG multi-doc retrieval works. No explicit aggregation UX. UIB-88 pending."),
        "UC-06": ("Done", "Session memory + TTL + isolation. 12 tests."),
        "UC-07": ("Done", "Session expiry + fresh context. 4 tests."),
        "UC-08": ("Partial", "Prompt-based clarification only. 3 tests."),
        "UC-09": ("Done", "Fallback detection + safe response. 11 tests."),
        "UC-10": ("Done", "Escalation store + Slack webhook. 13 tests."),
        "UC-11": ("Done", "Form guidance + location guidance. 11 tests. AC4 of UIB-143 blocked on UC-16."),
        "UC-12": ("Done", "Workflow guard + endpoint interception. 21 tests."),
        "UC-13": ("Done", "Prompt-based irrelevant query handling. 3 tests."),
        "UC-14": ("Done", "Two-layer guardrails (regex + LLM). 27 tests."),
        "UC-15": ("Done", "Feedback store + rating + comments. 12 tests."),
        "UC-16": ("Partial", "Only escalation API exists. No admin UI for 9/10 features."),
    }

    for row in range(2, ws_uc.max_row + 1):
        uc_id = ws_uc.cell(row, 1).value  # Column A = UC ID
        if uc_id and uc_id in UC_STATUS:
            status, notes = UC_STATUS[uc_id]
            ws_uc.cell(row, 8).value = status     # Column H = Implementation Status
            ws_uc.cell(row, 9).value = notes       # Column I = Gap/Notes

    # ── Sheet: Sheet1 (compact story tracker) ─────────────────────────────
    ws3 = wb["Sheet1"]
    for row in range(2, ws3.max_row + 1):
        story_id = ws3.cell(row, 3).value  # Column C = Story ID
        if not story_id or story_id not in STORY_STATUS:
            continue
        status, notes = STORY_STATUS[story_id]
        ws3.cell(row, 5).value = status    # Column E = Implementation Status
        ws3.cell(row, 6).value = notes     # Column F = Gap/Notes

    wb.save(str(path))
    print(f"✅ Updated: {path.name}")
    print(f"   Story Tracker: {ws.max_row - 1} rows")
    print(f"   Internal Chatbot Requirements: {ws_uc.max_row - 1} rows")
    print(f"   Sheet1: {ws3.max_row - 1} rows")


def update_validation_xlsx():
    """Update docs/Validation_of_Internal_Chatbot_Use_cases.xlsx."""
    path = PROJECT_ROOT / "docs" / "Validation_of_Internal_Chatbot_Use_cases.xlsx"
    wb = openpyxl.load_workbook(str(path))
    ws = wb["Sheet1"]

    for row, (status, gaps, obs) in UC_VALIDATION.items():
        ws.cell(row, 2).value = status  # Column B = Implementation Status
        ws.cell(row, 3).value = gaps    # Column C = Validation Gaps
        ws.cell(row, 4).value = obs     # Column D = Observations and Remarks

    wb.save(str(path))
    print(f"✅ Updated: {path.name}")
    print(f"   {len(UC_VALIDATION)} UC rows updated (rows {min(UC_VALIDATION)}–{max(UC_VALIDATION)})")


def main():
    tag_counts = count_test_tags()
    total = sum(tag_counts.values())
    print(f"Found {total} AC tags across {len(tag_counts)} stories\n")

    update_requirements_xlsx(tag_counts)
    print()
    update_validation_xlsx()

    print(f"\n{'━' * 40}")
    print("Done. Review changes, then commit.")


if __name__ == "__main__":
    main()
